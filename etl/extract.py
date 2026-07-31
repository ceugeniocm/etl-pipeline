"""Leitura em streaming de planilhas Excel (FR-001, FR-002, NFR-001).

O leitor abre a planilha em modo somente leitura e devolve as linhas uma a
uma, de modo que o consumo de memória acompanha o tamanho do bloco, e não o
do arquivo (NFR-001).

Dois formatos são atendidos pela mesma interface:

- ``.xlsx``/``.xlsm`` por meio do ``openpyxl`` (``read_only=True``,
  ``data_only=True``);
- ``.xls`` legado por meio do ``xlrd`` (tarefa 21 de ``docs/tasks.md``).

Uso típico::

    with open_source(config.source) as sheet:
        for chunk in iter_chunks(sheet.rows(), config.source.chunk_size):
            ...

Toda falha de origem — arquivo ausente, formato não suportado, planilha
corrompida, aba inexistente, cabeçalho ausente ou aba vazia — é sinalizada
como :class:`~etl.errors.ExtractionError` com mensagem em ``pt_BR`` (FR-001,
NFR-007).
"""

from __future__ import annotations

import os
from collections.abc import Iterable, Iterator, Sequence
from dataclasses import dataclass
from typing import Any

from etl import messages
from etl.config import DEFAULT_CHUNK_SIZE, DEFAULT_HEADER_ROW, SourceConfig
from etl.errors import ExtractionError
from etl.logging_setup import get_logger

__all__ = [
    "XLSX_EXTENSIONS",
    "XLS_EXTENSIONS",
    "SUPPORTED_EXTENSIONS",
    "SourceRow",
    "SheetReader",
    "open_sheet",
    "open_source",
    "iter_chunks",
]

_logger = get_logger(__name__)

#: Extensões lidas pelo ``openpyxl``.
XLSX_EXTENSIONS = (".xlsx", ".xlsm")

#: Extensões lidas pelo ``xlrd`` (formato legado).
XLS_EXTENSIONS = (".xls",)

#: Todas as extensões aceitas como origem (FR-001).
SUPPORTED_EXTENSIONS = XLSX_EXTENSIONS + XLS_EXTENSIONS


# --------------------------------------------------------------------------
# Linha de origem (tarefa 17 de docs/tasks.md)
# --------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class SourceRow:
    """Uma linha de dados da planilha, com sua procedência (NFR-009).

    :param sheet: nome da aba de onde a linha foi lida.
    :param number: número da linha **na planilha**, começando em 1, o mesmo
        que o usuário vê no Excel; é o número citado nas rejeições (FR-006).
    :param values: valores da linha indexados pelo nome da coluna de origem.
    """

    sheet: str
    number: int
    values: dict[str, Any]

    def __getitem__(self, column: str) -> Any:
        """Valor da coluna ``column``."""
        return self.values[column]

    def get(self, column: str, default: Any = None) -> Any:
        """Valor da coluna ``column`` ou ``default`` se ela não existir."""
        return self.values.get(column, default)


# --------------------------------------------------------------------------
# Interface comum aos formatos (tarefas 15, 16 e 21)
# --------------------------------------------------------------------------


class SheetReader:
    """Aba aberta para leitura, com cabeçalho já derivado.

    Instâncias são criadas por :func:`open_sheet`/:func:`open_source` e devem
    ser fechadas, preferencialmente pelo gerenciador de contexto. As linhas
    são produzidas sob demanda: :meth:`rows` é um gerador de passagem única.
    """

    def __init__(self, path: str, sheet: str, columns: Sequence[str]) -> None:
        self.path = path
        #: Nome da aba efetivamente lida.
        self.sheet = sheet
        #: Nomes das colunas do cabeçalho, na ordem da planilha.
        self.columns: tuple[str, ...] = tuple(columns)
        self._consumed = False
        self._closed = False

    # -- Ciclo de vida ----------------------------------------------------

    def __enter__(self) -> "SheetReader":
        """Devolve o próprio leitor, para uso com ``with``."""
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        """Libera os recursos da planilha ao sair do bloco ``with``."""
        self.close()

    def close(self) -> None:
        """Fecha a planilha; chamadas repetidas não têm efeito."""
        if not self._closed:
            self._closed = True
            self._close()

    # -- Leitura ----------------------------------------------------------

    def rows(self) -> Iterator[SourceRow]:
        """Itera as linhas de dados posteriores ao cabeçalho.

        Linhas totalmente vazias são ignoradas — planilhas costumam trazer
        linhas em branco no fim — sem interromper a leitura nem consumir a
        numeração de origem.

        :raises ExtractionError: se a aba não contiver nenhuma linha de dados
            (FR-001) ou se a leitura falhar no meio do arquivo.
        """
        if self._consumed:
            raise RuntimeError(
                "As linhas desta aba já foram lidas; abra a planilha novamente."
            )
        self._consumed = True

        count = 0
        width = len(self.columns)
        for number, values in self._iter_data_rows():
            if all(value is None for value in values):
                continue
            count += 1
            # Planilhas costumam trazer linhas mais curtas ou mais longas que
            # o cabeçalho; a linha é ajustada para que toda coluna do
            # cabeçalho exista, ainda que sem valor.
            if len(values) != width:
                values = tuple(values)[:width] + (None,) * (width - len(values))
            yield SourceRow(
                sheet=self.sheet,
                number=number,
                values=dict(zip(self.columns, values)),
            )

        if count == 0:
            raise ExtractionError(
                messages.ERR_SHEET_EMPTY.format(sheet=self.sheet)
            )
        _logger.debug(messages.INFO_ROWS_READ.format(sheet=self.sheet, count=count))

    # -- Pontos de extensão por formato -----------------------------------

    def _iter_data_rows(self) -> Iterator[tuple[int, Sequence[Any]]]:
        """Produz ``(número da linha, valores)`` já sem o cabeçalho."""
        raise NotImplementedError

    def _close(self) -> None:
        """Libera os recursos do formato concreto."""


# --------------------------------------------------------------------------
# Cabeçalho (tarefa 17)
# --------------------------------------------------------------------------


def _header_names(
    cells: Iterable[Any], *, sheet: str, path: str, header_row: int
) -> tuple[str, ...]:
    """Converte as células do cabeçalho em nomes de coluna.

    Células vazias recebem um nome reservado — elas não podem ser mapeadas,
    mas preservam a posição das colunas seguintes. Nomes repetidos tornariam
    o mapeamento ambíguo e por isso são recusados.

    :raises ExtractionError: se o cabeçalho não tiver nenhum nome utilizável
        ou se algum nome se repetir.
    """
    names: list[str] = []
    seen: set[str] = set()
    filled = 0

    for index, cell in enumerate(cells, start=1):
        name = "" if cell is None else str(cell).strip()
        if not name:
            names.append(f"__coluna_{index}__")
            continue
        if name in seen:
            raise ExtractionError(
                messages.ERR_HEADER_DUPLICATE_COLUMN.format(column=name, sheet=sheet)
            )
        seen.add(name)
        names.append(name)
        filled = index

    if not filled:
        raise ExtractionError(
            messages.ERR_HEADER_MISSING.format(sheet=sheet, row=header_row)
        )

    # Colunas vazias à direita do último cabeçalho preenchido não existem.
    del names[filled:]
    _logger.debug(
        messages.INFO_SHEET_SELECTED.format(sheet=sheet, path=path, columns=len(names))
    )
    return tuple(names)


# --------------------------------------------------------------------------
# Leitor .xlsx (tarefas 15 e 16)
# --------------------------------------------------------------------------


class _XlsxSheetReader(SheetReader):
    """Leitor ``openpyxl`` em modo somente leitura (NFR-001)."""

    def __init__(self, path: str, sheet: str | None, header_row: int) -> None:
        try:
            import openpyxl
        except ImportError as error:  # pragma: no cover - dependência declarada
            raise ExtractionError(
                messages.ERR_SOURCE_MISSING_DEPENDENCY.format(
                    extension=".xlsx", package="openpyxl"
                ),
                cause=error,
            ) from error

        try:
            self._workbook = openpyxl.load_workbook(
                path, read_only=True, data_only=True
            )
        except Exception as error:
            raise ExtractionError(
                messages.ERR_SOURCE_UNREADABLE.format(path=path), cause=error
            ) from error

        try:
            worksheet = self._select_worksheet(path, sheet)
            self._rows = worksheet.iter_rows(values_only=True)
            columns = _header_names(
                self._read_header(worksheet, path, header_row),
                sheet=worksheet.title,
                path=path,
                header_row=header_row,
            )
        except BaseException:
            self._workbook.close()
            raise

        super().__init__(path, worksheet.title, columns)
        self._next_number = header_row + 1

    def _select_worksheet(self, path: str, sheet: str | None):
        """Devolve a aba configurada ou, na ausência dela, a primeira."""
        if sheet is None:
            return self._workbook.worksheets[0]
        if sheet not in self._workbook.sheetnames:
            raise ExtractionError(
                messages.ERR_SHEET_NOT_FOUND.format(sheet=sheet, path=path)
            )
        return self._workbook[sheet]

    def _read_header(self, worksheet, path: str, header_row: int) -> Sequence[Any]:
        """Avança o iterador até o cabeçalho e devolve suas células.

        Uma aba sem nenhuma linha é relatada como vazia; uma aba que tenha
        linhas, mas nenhuma na posição do cabeçalho, é relatada como sem
        cabeçalho.
        """
        header = next(self._rows, None)
        if header is None:
            raise ExtractionError(
                messages.ERR_SHEET_EMPTY.format(sheet=worksheet.title)
            )
        for _ in range(header_row - 1):
            header = next(self._rows, None)
            if header is None:
                raise ExtractionError(
                    messages.ERR_HEADER_MISSING.format(
                        sheet=worksheet.title, row=header_row
                    )
                )
        return header

    def _iter_data_rows(self) -> Iterator[tuple[int, Sequence[Any]]]:
        """Itera as linhas restantes da aba, numeradas como na planilha."""
        number = self._next_number
        try:
            for values in self._rows:
                yield number, values
                number += 1
        except ExtractionError:
            raise
        except Exception as error:
            raise ExtractionError(
                messages.ERR_SOURCE_UNREADABLE.format(path=self.path), cause=error
            ) from error

    def _close(self) -> None:
        """Fecha a pasta de trabalho e o arquivo subjacente."""
        self._workbook.close()


# --------------------------------------------------------------------------
# Leitor .xls legado (tarefa 21)
# --------------------------------------------------------------------------


def _xls_cell_value(cell: Any, xlrd: Any, datemode: int) -> Any:
    """Converte uma célula do ``xlrd`` no tipo Python equivalente.

    Células vazias e células de erro viram ``None``; números de série de data
    viram ``datetime``; booleanos, inteiros do ``xlrd``, viram ``bool``. Assim
    ``.xls`` e ``.xlsx`` entregam os mesmos tipos ao restante do pipeline.
    """
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR):
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    return cell.value


class _XlsSheetReader(SheetReader):
    """Leitor ``xlrd`` para o formato legado ``.xls``.

    O ``xlrd`` devolve datas como números de série e booleanos como inteiros;
    a conversão é feita aqui para que os dois formatos entreguem os mesmos
    tipos ao restante do pipeline.
    """

    def __init__(self, path: str, sheet: str | None, header_row: int) -> None:
        try:
            import xlrd
        except ImportError as error:
            raise ExtractionError(
                messages.ERR_SOURCE_MISSING_DEPENDENCY.format(
                    extension=".xls", package="xlrd"
                ),
                cause=error,
            ) from error

        self._xlrd = xlrd
        try:
            self._book = xlrd.open_workbook(path, on_demand=True)
        except Exception as error:
            raise ExtractionError(
                messages.ERR_SOURCE_UNREADABLE.format(path=path), cause=error
            ) from error

        try:
            self._worksheet = self._select_worksheet(path, sheet)
            if self._worksheet.nrows == 0:
                raise ExtractionError(
                    messages.ERR_SHEET_EMPTY.format(sheet=self._worksheet.name)
                )
            if self._worksheet.nrows < header_row:
                raise ExtractionError(
                    messages.ERR_HEADER_MISSING.format(
                        sheet=self._worksheet.name, row=header_row
                    )
                )
            columns = _header_names(
                self._row_values(header_row - 1),
                sheet=self._worksheet.name,
                path=path,
                header_row=header_row,
            )
        except BaseException:
            self._book.release_resources()
            raise

        super().__init__(path, self._worksheet.name, columns)
        self._first_index = header_row

    def _select_worksheet(self, path: str, sheet: str | None):
        """Devolve a aba configurada ou, na ausência dela, a primeira."""
        if sheet is None:
            if not self._book.nsheets:
                raise ExtractionError(
                    messages.ERR_SOURCE_UNREADABLE.format(path=path)
                )
            return self._book.sheet_by_index(0)
        if sheet not in self._book.sheet_names():
            raise ExtractionError(
                messages.ERR_SHEET_NOT_FOUND.format(sheet=sheet, path=path)
            )
        return self._book.sheet_by_name(sheet)

    def _row_values(self, index: int) -> list[Any]:
        """Valores da linha ``index`` (base 0) já convertidos para tipos Python."""
        datemode = self._book.datemode
        return [
            _xls_cell_value(cell, self._xlrd, datemode)
            for cell in self._worksheet.row(index)
        ]

    def _iter_data_rows(self) -> Iterator[tuple[int, Sequence[Any]]]:
        """Itera as linhas posteriores ao cabeçalho, numeradas como na planilha."""
        try:
            for index in range(self._first_index, self._worksheet.nrows):
                yield index + 1, self._row_values(index)
        except Exception as error:
            raise ExtractionError(
                messages.ERR_SOURCE_UNREADABLE.format(path=self.path), cause=error
            ) from error

    def _close(self) -> None:
        """Libera os recursos mantidos pelo ``xlrd``."""
        self._book.release_resources()


# --------------------------------------------------------------------------
# API pública (tarefas 15, 16, 18 e 19)
# --------------------------------------------------------------------------


def open_sheet(
    path: str | os.PathLike[str],
    *,
    sheet: str | None = None,
    header_row: int = DEFAULT_HEADER_ROW,
) -> SheetReader:
    """Abre ``path`` e devolve a aba pronta para leitura.

    :param path: caminho da planilha de origem.
    :param sheet: nome da aba; ``None`` seleciona a primeira (FR-001).
    :param header_row: posição (base 1) da linha de cabeçalho.
    :raises ExtractionError: se o arquivo não existir, tiver extensão não
        suportada, estiver corrompido, não contiver a aba pedida ou não
        possuir cabeçalho.
    """
    path = os.fspath(path)
    extension = os.path.splitext(path)[1].lower()

    if extension not in SUPPORTED_EXTENSIONS:
        raise ExtractionError(
            messages.ERR_SOURCE_UNSUPPORTED_FORMAT.format(
                extension=extension or os.path.basename(path)
            )
        )
    if not os.path.exists(path):
        raise ExtractionError(messages.ERR_SOURCE_FILE_NOT_FOUND.format(path=path))
    if not os.path.isfile(path):
        raise ExtractionError(messages.ERR_SOURCE_UNREADABLE.format(path=path))

    if extension in XLS_EXTENSIONS:
        return _XlsSheetReader(path, sheet, header_row)
    return _XlsxSheetReader(path, sheet, header_row)


def open_source(source: SourceConfig) -> SheetReader:
    """Abre a planilha descrita pela seção ``source`` da configuração."""
    return open_sheet(
        source.path, sheet=source.sheet, header_row=source.header_row
    )


def iter_chunks(
    rows: Iterable[SourceRow], chunk_size: int = DEFAULT_CHUNK_SIZE
) -> Iterator[list[SourceRow]]:
    """Agrupa ``rows`` em listas de no máximo ``chunk_size`` linhas (FR-002).

    O último bloco pode ser menor; nenhum bloco vazio é produzido. Os blocos
    são montados sob demanda, mantendo em memória apenas o bloco corrente
    (NFR-001).

    :raises ValueError: se ``chunk_size`` não for positivo.
    """
    if chunk_size < 1:
        raise ValueError("chunk_size deve ser maior ou igual a 1")

    chunk: list[SourceRow] = []
    for row in rows:
        chunk.append(row)
        if len(chunk) == chunk_size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk
